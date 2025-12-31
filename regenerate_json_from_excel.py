#!/usr/bin/env python3
"""
Regenerate sample_data.json from SLA_Monthly_Status_Summary_FINAL.xlsx
Ensures Target values are preserved exactly as they appear in Excel (e.g., 10%, 20%, 0.5)
"""

import openpyxl
import json
from datetime import datetime

def parse_excel_to_json(excel_file):
    """Parse the Excel file and generate JSON matching the dashboard structure."""
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active
    
    data = []
    headers = []
    
    # Read header row (row 1)
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"Headers found: {headers}")
    print(f"Total columns: {len(headers)}")
    
    # Read data rows (starting from row 2)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        row_dict = {}
        for col_idx, (header, value) in enumerate(zip(headers, row)):
            if header is None:
                continue
                
            # Handle Target column specially - preserve EXACT format from Excel
            if header == "Target":
                if value is None or value == "":
                    row_dict[header] = ""
                elif isinstance(value, (int, float)):
                    # Check if it's a percentage (0.01 to 1.0 range typically means percentage)
                    # BUT we preserve the EXACT value from Excel
                    # Excel stores 10% as 0.1, 20% as 0.2, etc.
                    # Dashboard will display exactly as stored
                    row_dict[header] = value
                else:
                    # String values like "5 Days", "NA", etc.
                    row_dict[header] = str(value).strip()
            else:
                # Handle other columns
                if value is None:
                    row_dict[header] = ""
                elif isinstance(value, datetime):
                    row_dict[header] = value.strftime('%Y-%m-%d')
                elif isinstance(value, (int, float)):
                    row_dict[header] = value
                else:
                    row_dict[header] = str(value).strip()
        
        # Only add rows with at least a Project Name
        if row_dict.get("Project Name"):
            data.append(row_dict)
    
    print(f"\nTotal rows processed: {len(data)}")
    
    # Find and display Maruti Suzuki Diversity entries for verification
    print("\n" + "="*80)
    print("VERIFICATION: Maruti Suzuki Diversity Entries")
    print("="*80)
    maruti_diversity_count = 0
    for entry in data:
        if "Maruti Suzuki" in entry.get("Project Name", "") and "Diversity" in entry.get("Performance Measure", ""):
            maruti_diversity_count += 1
            print(f"\nEntry {maruti_diversity_count}:")
            print(f"  Project: {entry.get('Project Name')}")
            print(f"  Measure: {entry.get('Performance Measure')}")
            print(f"  Target: {entry.get('Target')} (Type: {type(entry.get('Target')).__name__})")
            print(f"  Region: {entry.get('Region')}")
            print(f"  FY: {entry.get('FY')}")
    
    print(f"\nTotal Maruti Suzuki Diversity entries found: {maruti_diversity_count}")
    print("="*80)
    
    return data

def main():
    excel_file = 'SLA_Monthly_Status_Summary_FINAL.xlsx'
    output_file = 'sample_data.json'
    backup_file = 'sample_data.json.backup'
    
    print(f"Reading Excel file: {excel_file}")
    data = parse_excel_to_json(excel_file)
    
    # Backup existing JSON
    import shutil
    try:
        shutil.copy(output_file, backup_file)
        print(f"\n✓ Backed up existing {output_file} to {backup_file}")
    except FileNotFoundError:
        print(f"\n! No existing {output_file} to backup")
    
    # Write new JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"\n✓ Generated {output_file} with {len(data)} entries")
    
    # Display sample entries
    print("\n" + "="*80)
    print("SAMPLE ENTRIES (First 3)")
    print("="*80)
    for i, entry in enumerate(data[:3], 1):
        print(f"\nEntry {i}:")
        print(f"  Project: {entry.get('Project Name')}")
        print(f"  Measure: {entry.get('Performance Measure')}")
        print(f"  Target: {entry.get('Target')} (Type: {type(entry.get('Target')).__name__})")
        print(f"  Region: {entry.get('Region')}")
    
    print("\n" + "="*80)
    print("JSON REGENERATION COMPLETE")
    print("="*80)
    print(f"\n✓ Total entries: {len(data)}")
    print(f"✓ Output file: {output_file}")
    print(f"✓ Backup file: {backup_file}")
    print("\nNext step: Restart dashboard to load new data")

if __name__ == "__main__":
    main()
