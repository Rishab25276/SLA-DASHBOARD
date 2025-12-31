#!/usr/bin/env python3
"""
Complete JSON regeneration from SLA_Monthly_Status_Summary_FINAL.xlsx
Processes all 5 sheets and maintains exact data structure for the dashboard
"""

import openpyxl
import json
from datetime import datetime

def process_metrics_details_sheet(ws):
    """Process FY 25-26 Metrics Details sheet."""
    data = []
    headers = [cell.value for cell in ws[1]]
    
    print(f"\nProcessing: FY 25-26 Metrics Details")
    print(f"Headers: {headers[:10]}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        row_dict = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            
            # Normalize header names to match dashboard expectations
            header_normalized = header.strip() if isinstance(header, str) else header
            
            # Map Excel headers to dashboard JSON keys
            header_map = {
                'Project': 'Project Name',
                'Performance Measure ': 'Performance Measure',  # Note the space
                'April25 Score': 'Apr\'25',
                'May25 Score': 'May\'25',
                'June25 Score': 'Jun\'25',
                'July25 Score': 'Jul\'25',
                'Aug25 Score': 'Aug\'25',
                'Sep25 Score': 'Sep\'25',
                'Oct25 Score': 'Oct\'25',
                'April MET/NOT_MET': 'Apr MET/NOT_MET',
                'May MET/NOT_MET': 'May MET/NOT_MET',
                'June MET/NOT_MET': 'Jun MET/NOT_MET',
                'July MET/NOT_MET': 'Jul MET/NOT_MET',
                'Aug MET/NOT_MET': 'Aug MET/NOT_MET',
                'Sep MET/NOT_MET': 'Sep MET/NOT_MET',
                'Oct MET/NOT_MET': 'Oct MET/NOT_MET'
            }
            
            json_key = header_map.get(header_normalized, header_normalized)
            
            # Handle values
            if value is None:
                row_dict[json_key] = ""
            elif isinstance(value, datetime):
                row_dict[json_key] = value.strftime('%Y-%m-%d')
            elif isinstance(value, (int, float)):
                row_dict[json_key] = value
            else:
                row_dict[json_key] = str(value).strip()
        
        # Add FY field
        row_dict['FY'] = 'FY 25-26'
        
        if row_dict.get("Project Name"):
            data.append(row_dict)
    
    print(f"  Rows processed: {len(data)}")
    return data

def process_summary_sheet(ws, fy_label):
    """Process FY Summary sheets."""
    data = []
    headers = [cell.value for cell in ws[1]]
    
    print(f"\nProcessing: {fy_label} Summary")
    print(f"Headers: {headers[:10]}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        row_dict = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            
            header_normalized = header.strip() if isinstance(header, str) else header
            
            # Map to dashboard keys
            header_map = {
                'Project': 'Project Name',
                'Regional Head': 'Regional Head',
                'Regional Head ': 'Regional Head'  # Handle trailing space
            }
            
            json_key = header_map.get(header_normalized, header_normalized)
            
            if value is None:
                row_dict[json_key] = ""
            elif isinstance(value, (int, float)):
                row_dict[json_key] = value
            else:
                row_dict[json_key] = str(value).strip()
        
        row_dict['FY'] = fy_label
        
        if row_dict.get("Project Name"):
            data.append(row_dict)
    
    print(f"  Rows processed: {len(data)}")
    return data

def process_not_reported_sheet(ws, fy_label):
    """Process Not Reported sheets."""
    data = []
    headers = [cell.value for cell in ws[1]]
    
    print(f"\nProcessing: {fy_label} Not Reported")
    print(f"Headers: {headers[:10]}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        row_dict = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            
            header_normalized = header.strip() if isinstance(header, str) else header
            
            # Map to dashboard keys
            header_map = {
                'Project': 'Project Name',
                'Regional Head': 'Regional Head',
                'Regional Head ': 'Regional Head'
            }
            
            json_key = header_map.get(header_normalized, header_normalized)
            
            if value is None:
                row_dict[json_key] = ""
            elif isinstance(value, (int, float)):
                row_dict[json_key] = value
            else:
                row_dict[json_key] = str(value).strip()
        
        row_dict['FY'] = fy_label
        
        if row_dict.get("Project Name"):
            data.append(row_dict)
    
    print(f"  Rows processed: {len(data)}")
    return data

def main():
    excel_file = 'SLA_Monthly_Status_Summary_FINAL.xlsx'
    output_file = 'sample_data.json'
    backup_file = 'sample_data.json.backup'
    
    print(f"{'='*80}")
    print(f"JSON REGENERATION FROM EXCEL")
    print(f"{'='*80}")
    print(f"Excel file: {excel_file}")
    
    # Backup existing JSON
    import shutil
    try:
        shutil.copy(output_file, backup_file)
        print(f"\n✓ Backed up {output_file} to {backup_file}")
    except FileNotFoundError:
        print(f"\n! No existing {output_file} to backup")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    print(f"\nSheet names: {wb.sheetnames}")
    
    # Process all sheets
    all_data = []
    
    # 1. FY 25-26 Metrics Details (main data source)
    metrics_ws = wb['FY 25-26 Metrics Details ']
    metrics_data = process_metrics_details_sheet(metrics_ws)
    all_data.extend(metrics_data)
    
    # 2. FY 24-25 Summary
    fy2425_summary_ws = wb['FY 24-25 Summary']
    fy2425_summary_data = process_summary_sheet(fy2425_summary_ws, 'FY 24-25')
    all_data.extend(fy2425_summary_data)
    
    # 3. FY 25-26 Summary
    fy2526_summary_ws = wb['FY 25-26 Summary']
    fy2526_summary_data = process_summary_sheet(fy2526_summary_ws, 'FY 25-26')
    all_data.extend(fy2526_summary_data)
    
    # 4. FY24-25 Not Reported
    fy2425_nr_ws = wb['FY24-25 Not Reported']
    fy2425_nr_data = process_not_reported_sheet(fy2425_nr_ws, 'FY 24-25')
    all_data.extend(fy2425_nr_data)
    
    # 5. FY25-26 Not Reported
    fy2526_nr_ws = wb['FY25-26 Not Reported']
    fy2526_nr_data = process_not_reported_sheet(fy2526_nr_ws, 'FY 25-26')
    all_data.extend(fy2526_nr_data)
    
    # Write JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*80}")
    print(f"JSON GENERATION COMPLETE")
    print(f"{'='*80}")
    print(f"✓ Total entries: {len(all_data)}")
    print(f"✓ Output file: {output_file}")
    print(f"✓ Backup file: {backup_file}")
    
    # Verify Maruti Suzuki Diversity
    print(f"\n{'='*80}")
    print(f"VERIFICATION: Maruti Suzuki Diversity Entry")
    print(f"{'='*80}")
    
    for entry in all_data:
        project = entry.get('Project Name', '')
        measure = entry.get('Performance Measure', '')
        if 'Maruti' in project and 'diversity' in measure.lower():
            print(f"\n✓ FOUND Maruti Suzuki Diversity:")
            print(f"  Project: {project}")
            print(f"  Measure: {measure}")
            print(f"  Target: {entry.get('Target')}")
            print(f"  Target Type: {type(entry.get('Target')).__name__}")
            if isinstance(entry.get('Target'), float):
                print(f"  Target as %: {entry.get('Target') * 100}%")
            print(f"  Region: {entry.get('Region')}")
            print(f"  FY: {entry.get('FY')}")
    
    print(f"\n{'='*80}")
    print(f"Ready to restart dashboard with new data!")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()
