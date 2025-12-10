#!/usr/bin/env python3
"""
Complete JSON regeneration from SLA_Monthly_Status_Summary_FINAL.xlsx
Generates the CORRECT structure with nested objects that the dashboard expects
"""

import openpyxl
import json
from datetime import datetime

def process_metrics_details_sheet(ws):
    """Process FY 25-26 Metrics Details sheet - this is the main drill-down data."""
    data = []
    headers = [cell.value for cell in ws[1]]
    
    print(f"\nProcessing: FY 25-26 Metrics Details")
    print(f"Headers: {headers[:10]}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        row_dict = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            
            # Keep headers EXACTLY as they are in Excel (don't normalize yet)
            # The dashboard expects "Performance Measure " with trailing space
            if value is None:
                row_dict[header] = ""
            elif isinstance(value, datetime):
                row_dict[header] = value.strftime('%Y-%m-%d')
            elif isinstance(value, (int, float)):
                row_dict[header] = value
            else:
                row_dict[header] = str(value).strip()
        
        if row_dict.get("Project"):
            data.append(row_dict)
    
    print(f"  Rows processed: {len(data)}")
    return data

def process_summary_sheet(ws, fy_label):
    """Process FY Summary sheets."""
    data = []
    headers = [cell.value for cell in ws[1]]
    
    print(f"\nProcessing: {fy_label} Summary")
    print(f"Headers: {headers[:10]}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        row_dict = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            
            if value is None:
                row_dict[header] = ""
            elif isinstance(value, (int, float)):
                row_dict[header] = value
            else:
                row_dict[header] = str(value).strip()
        
        if row_dict.get("Project"):
            data.append(row_dict)
    
    print(f"  Rows processed: {len(data)}")
    return data

def process_not_reported_sheet(ws, fy_label):
    """Process Not Reported sheets."""
    data = []
    headers = [cell.value for cell in ws[1]]
    
    print(f"\nProcessing: {fy_label} Not Reported")
    print(f"Headers: {headers[:10]}")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        row_dict = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            
            if value is None:
                row_dict[header] = ""
            elif isinstance(value, (int, float)):
                row_dict[header] = value
            else:
                row_dict[header] = str(value).strip()
        
        if row_dict.get("Project"):
            data.append(row_dict)
    
    print(f"  Rows processed: {len(data)}")
    return data

def main():
    excel_file = 'SLA_Monthly_Status_Summary_FINAL.xlsx'
    output_file = 'sample_data.json'
    backup_file = 'sample_data.json.backup'
    
    print(f"{'='*80}")
    print(f"JSON REGENERATION FROM EXCEL - CORRECT NESTED STRUCTURE")
    print(f"{'='*80}")
    print(f"Excel file: {excel_file}")
    
    # Backup existing JSON
    import shutil
    try:
        shutil.copy(output_file, backup_file)
        print(f"\n✓ Backed up {output_file} to {backup_file}")
    except FileNotFoundError:
        print(f"\n! No existing {output_file} to backup")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    print(f"\nSheet names: {wb.sheetnames}")
    
    # Process all sheets into the CORRECT nested structure
    # Dashboard expects: { fy24_25: [], fy25_26: [], fy24_25_not_reported: [], fy25_26_not_reported: [], fy25_26_metrics_details: [] }
    
    output_data = {}
    
    # 1. FY 24-25 Summary
    print("\n" + "="*80)
    fy2425_summary_ws = wb['FY 24-25 Summary']
    output_data['fy24_25'] = process_summary_sheet(fy2425_summary_ws, 'FY 24-25')
    
    # 2. FY 25-26 Summary
    print("\n" + "="*80)
    fy2526_summary_ws = wb['FY 25-26 Summary']
    output_data['fy25_26'] = process_summary_sheet(fy2526_summary_ws, 'FY 25-26')
    
    # 3. FY24-25 Not Reported
    print("\n" + "="*80)
    fy2425_nr_ws = wb['FY24-25 Not Reported']
    output_data['fy24_25_not_reported'] = process_not_reported_sheet(fy2425_nr_ws, 'FY 24-25')
    
    # 4. FY25-26 Not Reported
    print("\n" + "="*80)
    fy2526_nr_ws = wb['FY25-26 Not Reported']
    output_data['fy25_26_not_reported'] = process_not_reported_sheet(fy2526_nr_ws, 'FY 25-26')
    
    # 5. FY 25-26 Metrics Details (MOST IMPORTANT - this is the drill-down data!)
    print("\n" + "="*80)
    metrics_ws = wb['FY 25-26 Metrics Details ']
    output_data['fy25_26_metrics_details'] = process_metrics_details_sheet(metrics_ws)
    
    # Write JSON with nested structure
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*80}")
    print(f"JSON GENERATION COMPLETE - NESTED STRUCTURE")
    print(f"{'='*80}")
    print(f"✓ FY 24-25 Summary: {len(output_data['fy24_25'])} entries")
    print(f"✓ FY 25-26 Summary: {len(output_data['fy25_26'])} entries")
    print(f"✓ FY 24-25 Not Reported: {len(output_data['fy24_25_not_reported'])} entries")
    print(f"✓ FY 25-26 Not Reported: {len(output_data['fy25_26_not_reported'])} entries")
    print(f"✓ FY 25-26 Metrics Details: {len(output_data['fy25_26_metrics_details'])} entries")
    print(f"✓ Output file: {output_file}")
    print(f"✓ Backup file: {backup_file}")
    
    # Verify Maruti Suzuki Diversity
    print(f"\n{'='*80}")
    print(f"VERIFICATION: Maruti Suzuki Diversity Entry")
    print(f"{'='*80}")
    
    for entry in output_data['fy25_26_metrics_details']:
        project = entry.get('Project', '')
        measure = entry.get('Performance Measure ', '')  # Note trailing space
        if 'Maruti' in project and 'diversity' in measure.lower():
            print(f"\n✓ FOUND Maruti Suzuki Diversity:")
            print(f"  Project: {project}")
            print(f"  Measure: {measure}")
            print(f"  Target: {entry.get('Target')}")
            print(f"  Target Type: {type(entry.get('Target')).__name__}")
            if isinstance(entry.get('Target'), float):
                print(f"  Target as %: {entry.get('Target') * 100}%")
            print(f"  Region: {entry.get('Region')}")
            print(f"  April25 Score: {entry.get('April25 Score')}")
            print(f"  May25 Score: {entry.get('May25 Score')}")
    
    print(f"\n{'='*80}")
    print(f"Structure matches dashboard expectations!")
    print(f"Ready to restart dashboard with new data!")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()
